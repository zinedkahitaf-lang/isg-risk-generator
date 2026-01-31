import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import openai
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

class ISGGeneratorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Tosyalı İSG Risk Değerlendirme Asistanı (AI Powered)")
        self.root.geometry("600x450")
        
        # Styles
        style = ttk.Style()
        style.configure("TLabel", font=("Arial", 11))
        style.configure("TButton", font=("Arial", 11, "bold"))

        # Frame
        main_frame = ttk.Frame(root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Title
        title_label = ttk.Label(main_frame, text="İSG Risk Değerlendirme Oluşturucu", font=("Arial", 16, "bold"))
        title_label.pack(pady=(0, 20))

        # API Key Input
        ttk.Label(main_frame, text="OpenAI API Key:").pack(anchor=tk.W)
        self.api_key_entry = ttk.Entry(main_frame, width=50, show="*")
        self.api_key_entry.pack(fill=tk.X, pady=(0, 10))

        # Workplace Type Input
        ttk.Label(main_frame, text="Çalışma Alanı / Faaliyet Türü (Örn: Asansör Montajı):").pack(anchor=tk.W)
        self.workplace_entry = ttk.Entry(main_frame, width=50)
        self.workplace_entry.pack(fill=tk.X, pady=(0, 20))

        # Generate Button
        self.generate_btn = ttk.Button(main_frame, text="Risk Analizi Oluştur ve Kaydet", command=self.start_generation)
        self.generate_btn.pack(fill=tk.X, pady=10)

        # Log Area
        ttk.Label(main_frame, text="Durum:").pack(anchor=tk.W)
        self.log_text = tk.Text(main_frame, height=10, width=60, font=("Consolas", 9))
        self.log_text.pack(fill=tk.BOTH, expand=True)

    def log(self, message):
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)
        self.root.update()

    def start_generation(self):
        api_key = self.api_key_entry.get().strip()
        workplace = self.workplace_entry.get().strip()

        if not api_key:
            messagebox.showerror("Hata", "Lütfen OpenAI API Key giriniz.")
            return
        if not workplace:
            messagebox.showerror("Hata", "Lütfen bir çalışma alanı giriniz.")
            return

        self.generate_btn.config(state=tk.DISABLED)
        self.log("İşlem başlatılıyor...")
        
        try:
            data = self.fetch_risks_from_openai(api_key, workplace)
            if data:
                filename = f"ISG_Risk_Analizi_{workplace.replace(' ', '_')}.xlsx"
                self.create_excel(data, filename)
                self.log(f"BAŞARILI: '{filename}' oluşturuldu.")
                messagebox.showinfo("Başarılı", f"Dosya oluşturuldu:\n{filename}")
            else:
                self.log("HATA: AI verisi alınamadı.")
        except Exception as e:
            self.log(f"BEKLENMEYEN HATA: {str(e)}")
            messagebox.showerror("Hata", str(e))
        finally:
            self.generate_btn.config(state=tk.NORMAL)

    def fetch_risks_from_openai(self, api_key, workplace):
        self.log("OpenAI API'ye bağlanılıyor...")
        client = openai.OpenAI(api_key=api_key)
        
        prompt = f"""
        Sen uzman bir İSG (İş Sağlığı ve Güvenliği) mühendisisin.
        Görev: '{workplace}' işi için 20 adet detaylı risk değerlendirmesi yap.
        
        Çıktı formatı: Sadece saf JSON array döndür. Markdown bloğu kullanma.
        Objeleri şu anahtarlarla oluştur:
        - sira_no (1'den 20'ye kadar)
        - faaliyet_alani (Örn: Kaynak İşleri, Yüksekte Çalışma, vb.)
        - tehlike_tanimi
        - risk_tanimi
        - olasilik (1-10 arası tam sayı)
        - frekans (1-10 arası tam sayı)
        - siddet (1-100 arası tam sayı)
        - onlemler (DÖF - Detaylı önlemler maddeler halinde)
        - sonraki_olasilik (Önlem sonrası düşürülmüş 1-10 arası)
        - sonraki_frekans (Önlem sonrası düşürülmüş 1-10 arası)
        - sonraki_siddet (Önlem sonrası düşürülmüş 1-100 arası)

        ÖNEMLİ KURALLAR:
        1. Mevcut Skor (olasilik * frekans * siddet) hesaplamasını ben yapacağım, sen sadece değerleri ver.
        2. Sonraki Skor (sonraki_olasilik * sonraki_frekans * sonraki_siddet) değerinin KESİNLİKLE 70'in altında (Kabul Edilebilir Risk) olmasını sağla. Ona göre Sonraki değerleri düşür.
        3. 'Tosyalı' formatına uygun, profesyonel ve teknik bir dil kullan.
        """

        try:
            response = client.chat.completions.create(
                model="gpt-4o",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.7
            )
            content = response.choices[0].message.content.strip()
            # Clean possible markdown format
            if content.startswith("```json"):
                content = content[7:]
            if content.endswith("```"):
                content = content[:-3]
            
            data = json.loads(content)
            self.log("Veri başarıyla alındı ve işlendi.")
            return data
        except Exception as e:
            self.log(f"OpenAI Hatası: {str(e)}")
            return None

    def create_excel(self, risk_data, filename):
        self.log("Excel dosyası hazırlanıyor...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Risk Değerlendirme"

        # Headers
        headers = [
            "Sıra No", "Faaliyet Alanı", "Tehlike Tanımı", "Risk Tanımı", 
            "Olasılık", "Frekans", "Şiddet", "Mevcut Skor", 
            "Önlemler (DÖF)", 
            "Sonraki Olasılık", "Sonraki Frekans", "Sonraki Şiddet", "Sonraki Skor"
        ]
        
        ws.append(headers)

        # Style Headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add Data
        for item in risk_data:
            p = item.get('olasilik', 1)
            f = item.get('frekans', 1)
            s = item.get('siddet', 1)
            current_score = p * f * s

            np = item.get('sonraki_olasilik', 1)
            nf = item.get('sonraki_frekans', 1)
            ns = item.get('sonraki_siddet', 1)
            next_score = np * nf * ns
            
            row = [
                item.get('sira_no'),
                item.get('faaliyet_alani'),
                item.get('tehlike_tanimi'),
                item.get('risk_tanimi'),
                p, f, s, current_score,
                item.get('onlemler'),
                np, nf, ns, next_score
            ]
            ws.append(row)

        # Formatting and Conditions
        score_col_idx = 8  # 'Mevcut Skor' column index (1-based: H)
        next_score_col_idx = 13 # 'Sonraki Skor' column index (1-based: M)
        
        # Color definitions
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        # Apply styles and conditions row by row
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            # Border for all cells
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Conditional Formatting for Current Score directly
            c_score_cell = row[score_col_idx - 1]
            try:
                val = float(c_score_cell.value)
                if val >= 200:
                    c_score_cell.fill = red_fill
                elif 70 <= val < 200:
                    c_score_cell.fill = yellow_fill
                else:
                    c_score_cell.fill = green_fill
            except:
                pass

            # Conditional Formatting for Next Score
            n_score_cell = row[next_score_col_idx - 1]
            try:
                val = float(n_score_cell.value)
                if val >= 200:
                    n_score_cell.fill = red_fill
                elif 70 <= val < 200:
                    n_score_cell.fill = yellow_fill
                else:
                    n_score_cell.fill = green_fill
            except:
                pass

        # Column Widths
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['I'].width = 50
        
        wb.save(filename)

if __name__ == "__main__":
    root = tk.Tk()
    app = ISGGeneratorApp(root)
    root.mainloop()
